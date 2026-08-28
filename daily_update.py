from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, timedelta
import subprocess

# --------------------------------------------------
# SETTINGS
# --------------------------------------------------

REPO = Path(__file__).resolve().parent

# Change this if your actual Excel filename is different
EXCEL_FILE = REPO / "12603644.xlsx"

SHEET_NAME = "Daily Log"
LOG_FOLDER = REPO / "logs"


# --------------------------------------------------
# HELPER FUNCTIONS
# --------------------------------------------------

def ask_number(question):
    while True:
        try:
            value = input(question).strip()
            return int(value)
        except ValueError:
            print("Please enter a number.")


def ask_choice(question, choices):
    print()
    print(question)

    for i, choice in enumerate(choices, 1):
        print(f"  {i}. {choice}")

    while True:
        try:
            number = int(input("Choose: "))
            if 1 <= number <= len(choices):
                return choices[number - 1]
        except ValueError:
            pass

        print("Please choose a valid number.")


# --------------------------------------------------
# LOAD EXCEL
# --------------------------------------------------

if not EXCEL_FILE.exists():
    print(f"\nERROR: Excel file not found:")
    print(EXCEL_FILE)
    input("\nPress Enter to exit...")
    raise SystemExit

workbook = load_workbook(EXCEL_FILE)
sheet = workbook[SHEET_NAME]


# --------------------------------------------------
# FIND LAST ENTRY
# --------------------------------------------------

last_date = None
last_row = None

for row in range(6, sheet.max_row + 1):
    value = sheet.cell(row, 1).value

    if isinstance(value, datetime):
        last_date = value
        last_row = row

    elif value:
        try:
            last_date = datetime.strptime(str(value), "%Y-%m-%d")
            last_row = row
        except ValueError:
            pass


if last_date:
    next_date = last_date + timedelta(days=1)
else:
    next_date = datetime.now()


# --------------------------------------------------
# SHOW ENTRY DATE
# --------------------------------------------------

print("\n========================================")
print("        DAILY TASK LOGGER")
print("========================================")

print(f"\nNext entry: {next_date.strftime('%d %B %Y')}")


# --------------------------------------------------
# COLLECT DATA
# --------------------------------------------------

print("\nEnter today's data.")
print("Durations are in MINUTES.\n")

sleep = ask_number("Sleep (minutes): ")
fitness = ask_number("Fitness (minutes): ")
study = ask_number("Study (minutes): ")
coding = ask_number("Coding (minutes): ")
class_time = ask_number("Class (minutes): ")
classes_attended = ask_number("Classes attended: ")
other = ask_number("Other activities (minutes): ")

feeling = ask_choice(
    "Today's feeling:",
    ["Low", "Neutral", "Good"]
)

satisfaction = ask_choice(
    "Satisfaction level:",
    ["Unsatisfied", "Neutral", "Satisfied"]
)

energy = ask_choice(
    "Energy level:",
    ["Low", "Medium", "High"]
)

notes = input("\nNotes: ").strip()


# --------------------------------------------------
# FIND EMPTY ROW
# --------------------------------------------------

target_row = None

for row in range(6, sheet.max_row + 1):
    if sheet.cell(row, 1).value is None:
        target_row = row
        break

if target_row is None:
    target_row = sheet.max_row + 1


# --------------------------------------------------
# WRITE TO EXCEL
# --------------------------------------------------

sheet.cell(target_row, 1).value = next_date
sheet.cell(target_row, 2).value = sleep
sheet.cell(target_row, 3).value = fitness
sheet.cell(target_row, 4).value = study
sheet.cell(target_row, 5).value = coding
sheet.cell(target_row, 6).value = class_time
sheet.cell(target_row, 7).value = classes_attended
sheet.cell(target_row, 8).value = other

# Preserve the formulas already used by your workbook
sheet.cell(target_row, 9).value = (
    f'=IF(SUM(B{target_row}:F{target_row},H{target_row})=0,"",'
    f'SUM(B{target_row}:F{target_row},H{target_row}))'
)

sheet.cell(target_row, 10).value = (
    f'=IF(I{target_row}="","",1440-I{target_row})'
)

sheet.cell(target_row, 11).value = feeling
sheet.cell(target_row, 12).value = satisfaction
sheet.cell(target_row, 13).value = energy
sheet.cell(target_row, 14).value = notes


# Copy formatting from previous row
if target_row > 6:
    source_row = target_row - 1

    for column in range(1, 15):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)

        if source.has_style:
            target._style = source._style

        if source.number_format:
            target.number_format = source.number_format


# Make sure date is displayed correctly
sheet.cell(target_row, 1).number_format = "dd-mmm-yyyy"


# Save Excel
workbook.save(EXCEL_FILE)


# --------------------------------------------------
# CREATE DAILY MARKDOWN LOG
# --------------------------------------------------

LOG_FOLDER.mkdir(exist_ok=True)

date_string = next_date.strftime("%Y-%m-%d")

log_file = LOG_FOLDER / f"{date_string}.md"

total_tracked = sleep + fitness + study + coding + class_time + other
free_time = 1440 - total_tracked

log_content = f"""# Daily Log — {next_date.strftime("%d %B %Y")}

## Time

| Activity | Minutes |
|---|---:|
| Sleep | {sleep} |
| Fitness | {fitness} |
| Study | {study} |
| Coding | {coding} |
| Class | {class_time} |
| Other Activities | {other} |
| **Total Tracked** | **{total_tracked}** |
| **Free / Unaccounted** | **{free_time}** |

## Academic

- Classes attended: **{classes_attended}**

## Reflection

- **Feeling:** {feeling}
- **Satisfaction:** {satisfaction}
- **Energy:** {energy}

## Notes

{notes if notes else "_No notes._"}
"""

log_file.write_text(log_content, encoding="utf-8")


# --------------------------------------------------
# GIT
# --------------------------------------------------

def run_git(command):
    result = subprocess.run(
        command,
        cwd=REPO,
        text=True,
        capture_output=True
    )

    if result.returncode != 0:
        print("\nGit error:")
        print(result.stderr)
        raise SystemExit

    return result.stdout.strip()


print("\n----------------------------------------")
print("Saving to GitHub...")
print("----------------------------------------")

run_git(["git", "add", "."])

commit_message = f"Daily log - {date_string}"

commit_result = subprocess.run(
    ["git", "commit", "-m", commit_message],
    cwd=REPO,
    text=True,
    capture_output=True
)

if commit_result.returncode != 0:
    print("\nGit commit failed:")
    print(commit_result.stdout)
    print(commit_result.stderr)

    input("\nPress Enter to exit...")
    raise SystemExit


push_result = subprocess.run(
    ["git", "push"],
    cwd=REPO,
    text=True,
    capture_output=True
)

if push_result.returncode != 0:
    print("\nGit push failed:")
    print(push_result.stdout)
    print(push_result.stderr)

    input("\nPress Enter to exit...")
    raise SystemExit


# --------------------------------------------------
# DONE
# --------------------------------------------------

print("\n========================================")
print("             SUCCESS")
print("========================================")

print(f"\nExcel updated:")
print(f"  {EXCEL_FILE.name}")

print(f"\nDaily log created:")
print(f"  logs/{date_string}.md")

print(f"\nGit commit:")
print(f"  {commit_message}")

print("\nEverything has been pushed to GitHub.")

input("\nPress Enter to exit...")